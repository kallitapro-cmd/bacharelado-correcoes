"""Testes unitários da exportação para Excel (Story 3.3).

Cobre as funções puras de excel_builder sem instanciar Streamlit.
AC-09: pelo menos 3 testes cobrindo os cenários especificados.
"""

from __future__ import annotations

from datetime import datetime

import pytest
from openpyxl import load_workbook
from packages.wrapper.schemas import FichaCorrecao

from src.batch.group_detector import GrupoCandidato
from src.batch.plagiarism_detector import ParPlagio
from src.ui.excel_builder import (
    construir_workbook,
    gerar_nome_arquivo,
    sanitizar_nome_arquivo,
    workbook_para_bytes,
)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _ra(i: int) -> str:
    return f"2026{i:07d}"


def _ficha(
    i: int,
    nota_calibrada: float = 7.0,
    nota_bruta: float | None = None,
) -> FichaCorrecao:
    return FichaCorrecao(
        ra=_ra(i),
        nota_a1=nota_calibrada,
        nota_a2=nota_bruta,
        feedback=f"Feedback aluno {i}",
        confianca="alta",
        flags=[],
    )


def _par_plagio(ra_a: str, ra_b: str, sim: float) -> ParPlagio:
    sev = "vermelho" if sim >= 0.90 else "amarelo"
    return ParPlagio(aluno_a=ra_a, aluno_b=ra_b, similaridade=sim, severidade=sev)


def _grupo(membros: list[str]) -> GrupoCandidato:
    return GrupoCandidato(
        membros=membros,
        confianca="media",  # type: ignore[arg-type]
        razao_confianca="Menção mútua.",
    )


def _decisao(acao: str = "aprovado", nota_final: float = 7.0, observacao: str = "") -> dict:
    return {
        "acao": acao,
        "nota_final": nota_final,
        "timestamp": datetime.now().isoformat(),
        "observacao": observacao,
    }


def _wb_para_ws(
    fichas: list[FichaCorrecao],
    decisoes: dict,
    alertas: list | None = None,
    grupos: list | None = None,
    nomes: dict | None = None,
    metadata: dict | None = None,
):
    wb = construir_workbook(
        fichas_calibradas=fichas,
        alertas_plagio=alertas or [],
        grupos_candidatos=grupos or [],
        decisoes=decisoes,
        nomes_por_ra=nomes or {},
        metadata=metadata or {},
    )
    return wb


# ---------------------------------------------------------------------------
# TC-01: geração do Excel com fichas mockadas produz arquivo com 2 abas (AC-09-a)
# ---------------------------------------------------------------------------


class TestConstruirWorkbook:
    def test_duas_abas_geradas(self) -> None:
        fichas = [_ficha(1), _ficha(2)]
        decisoes = {_ra(1): _decisao(), _ra(2): _decisao()}
        wb = _wb_para_ws(fichas, decisoes)

        assert "Notas" in wb.sheetnames
        assert "Alertas" in wb.sheetnames
        assert len(wb.sheetnames) == 2

    def test_aba_notas_tem_cabecalho_correto(self) -> None:
        fichas = [_ficha(1)]
        decisoes = {_ra(1): _decisao()}
        wb = _wb_para_ws(fichas, decisoes)

        ws_notas = wb["Notas"]
        cabecalhos = [cell.value for cell in ws_notas[1]]
        assert "RA" in cabecalhos
        assert "Nota Final" in cabecalhos
        assert "Revisão Manual" in cabecalhos
        assert "Observações PA" in cabecalhos

    def test_aba_notas_tem_linha_por_aluno(self) -> None:
        fichas = [_ficha(i) for i in range(1, 6)]
        decisoes = {_ra(i): _decisao() for i in range(1, 6)}
        wb = _wb_para_ws(fichas, decisoes)

        ws_notas = wb["Notas"]
        # 1 cabeçalho + 5 linhas de alunos
        assert ws_notas.max_row == 6

    def test_nota_final_vem_de_decisoes(self) -> None:
        ra = _ra(1)
        ficha = _ficha(1, nota_calibrada=7.0)
        decisoes = {ra: _decisao(acao="editado", nota_final=8.5)}
        wb = _wb_para_ws([ficha], decisoes)

        ws = wb["Notas"]
        # Linha 2 (após cabeçalho), coluna Nota Final (índice 6, base 1)
        cabecalhos = [cell.value for cell in ws[1]]
        idx_nota_final = cabecalhos.index("Nota Final") + 1
        valor = ws.cell(row=2, column=idx_nota_final).value
        assert float(valor) == pytest.approx(8.5)

    def test_serializa_para_bytes(self) -> None:
        fichas = [_ficha(1)]
        decisoes = {_ra(1): _decisao()}
        wb = _wb_para_ws(fichas, decisoes)
        raw = workbook_para_bytes(wb)

        assert isinstance(raw, bytes)
        assert len(raw) > 0
        # Verifica que os bytes são um xlsx válido relendo com openpyxl
        wb2 = load_workbook(filename=__import__("io").BytesIO(raw))
        assert "Notas" in wb2.sheetnames


# ---------------------------------------------------------------------------
# TC-02: aluno com revisao_manual=True aparece com flag correta (AC-09-b)
# ---------------------------------------------------------------------------


class TestRevisaoManual:
    def test_flag_revisao_manual_true_na_planilha(self) -> None:
        ra = _ra(1)
        ficha = _ficha(1, nota_calibrada=6.5)
        decisoes = {ra: _decisao(acao="revisao_manual", nota_final=6.5)}
        wb = _wb_para_ws([ficha], decisoes)

        ws = wb["Notas"]
        cabecalhos = [cell.value for cell in ws[1]]
        idx_rev = cabecalhos.index("Revisão Manual") + 1
        valor = ws.cell(row=2, column=idx_rev).value
        assert valor is True

    def test_flag_revisao_manual_false_para_aprovado(self) -> None:
        ra = _ra(1)
        ficha = _ficha(1, nota_calibrada=7.0)
        decisoes = {ra: _decisao(acao="aprovado", nota_final=7.0)}
        wb = _wb_para_ws([ficha], decisoes)

        ws = wb["Notas"]
        cabecalhos = [cell.value for cell in ws[1]]
        idx_rev = cabecalhos.index("Revisão Manual") + 1
        valor = ws.cell(row=2, column=idx_rev).value
        assert valor is False

    def test_linha_revisao_manual_tem_fill_amarelo(self) -> None:
        ra = _ra(1)
        ficha = _ficha(1)
        decisoes = {ra: _decisao(acao="revisao_manual")}
        wb = _wb_para_ws([ficha], decisoes)

        ws = wb["Notas"]
        # Linha 2 = primeiro aluno (após cabeçalho)
        fill = ws.cell(row=2, column=1).fill
        assert fill.fgColor.rgb == "FFFF00FF" or fill.fgColor.rgb.endswith("FFFF00")

    def test_observacao_exportada(self) -> None:
        ra = _ra(1)
        ficha = _ficha(1)
        decisoes = {ra: _decisao(acao="aprovado", observacao="Nota ajustada.")}
        wb = _wb_para_ws([ficha], decisoes)

        ws = wb["Notas"]
        cabecalhos = [cell.value for cell in ws[1]]
        idx_obs = cabecalhos.index("Observações PA") + 1
        valor = ws.cell(row=2, column=idx_obs).value
        assert valor == "Nota ajustada."


# ---------------------------------------------------------------------------
# TC-03: nome do arquivo segue padrão notas_{turma}_{data}.xlsx (AC-09-c)
# ---------------------------------------------------------------------------


class TestNomeArquivo:
    def test_nome_com_turma_e_data(self) -> None:
        data = datetime(2026, 5, 22)
        nome = gerar_nome_arquivo("CC2025A", data)
        assert nome == "notas_CC2025A_20260522.xlsx"

    def test_turma_vazia_usa_fallback(self) -> None:
        data = datetime(2026, 5, 22)
        nome = gerar_nome_arquivo("", data)
        assert nome == "notas_turma_20260522.xlsx"

    def test_turma_com_espacos_sanitizada(self) -> None:
        nome = gerar_nome_arquivo("Turma A / 2026", datetime(2026, 5, 22))
        assert " " not in nome
        assert "/" not in nome
        assert nome.startswith("notas_Turma_A")

    def test_sanitizar_nome_preserva_alfanumerico(self) -> None:
        assert sanitizar_nome_arquivo("CC2025A") == "CC2025A"
        assert sanitizar_nome_arquivo("turma-01") == "turma-01"
        assert sanitizar_nome_arquivo("turma_01") == "turma_01"

    def test_sanitizar_nome_remove_especiais(self) -> None:
        resultado = sanitizar_nome_arquivo("Turma A/B:2026!")
        assert "/" not in resultado
        assert ":" not in resultado
        assert "!" not in resultado


# ---------------------------------------------------------------------------
# TC-04: aba Alertas contém alunos com plágio/grupo (extra)
# ---------------------------------------------------------------------------


class TestAbaAlertas:
    def test_aba_alertas_tem_cabecalho(self) -> None:
        wb = _wb_para_ws([], {})
        ws = wb["Alertas"]
        cabecalhos = [cell.value for cell in ws[1]]
        assert "RA" in cabecalhos
        assert "Tipo de Alerta" in cabecalhos
        assert "Decisão do PA" in cabecalhos

    def test_aluno_com_plagio_aparece_em_alertas(self) -> None:
        ra1 = _ra(1)
        ra2 = _ra(2)
        fichas = [_ficha(1), _ficha(2)]
        par = _par_plagio(ra1, ra2, sim=0.80)
        decisoes = {ra1: _decisao(), ra2: _decisao()}
        wb = _wb_para_ws(fichas, decisoes, alertas=[par])

        ws = wb["Alertas"]
        ras_na_aba = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert ra1 in ras_na_aba
        assert ra2 in ras_na_aba

    def test_plagio_abaixo_limiar_nao_aparece_em_alertas(self) -> None:
        ra1 = _ra(1)
        ra2 = _ra(2)
        fichas = [_ficha(1), _ficha(2)]
        # similaridade 0.60 — abaixo do limiar moderado de 0.70
        par = _par_plagio(ra1, ra2, sim=0.60)
        decisoes = {ra1: _decisao(), ra2: _decisao()}
        wb = _wb_para_ws(fichas, decisoes, alertas=[par])

        ws = wb["Alertas"]
        # Apenas cabeçalho — nenhum aluno na aba
        assert ws.max_row == 1
