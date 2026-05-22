"""Módulo de geração do arquivo Excel para exportação (Story 3.3).

Lógica pura — sem dependência de Streamlit. Testável independentemente.

Consome:
- fichas_calibradas: list[FichaCorrecao] (de batch_results)
- alertas_plagio: list[ParPlagio] (de batch_results)
- grupos_candidatos: list[GrupoCandidato] (de batch_results)
- decisoes: dict[str, dict] (de st.session_state["decisoes"] — Story 3.2)
- nomes_por_ra: dict[str, str] (de st.session_state["nomes_por_ra"] — Story 3.0)
- metadata: dict (de batch_config["metadata"] — Story 3.0)

Produz:
- workbook openpyxl com abas "Notas" e "Alertas"
"""

from __future__ import annotations

import io
import re
from datetime import datetime
from typing import TYPE_CHECKING, Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

if TYPE_CHECKING:
    from packages.wrapper.schemas import FichaCorrecao

    from src.batch.group_detector import GrupoCandidato
    from src.batch.plagiarism_detector import ParPlagio

# Preenchimento amarelo para revisão manual (AC-05)
_FILL_REVISAO_MANUAL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_FILL_CABECALHO = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_FONT_CABECALHO = Font(color="FFFFFF", bold=True)

_LIMIAR_PLAGIO_SEVERO = 0.85
_LIMIAR_PLAGIO_MODERADO = 0.70


def sanitizar_nome_arquivo(turma: str) -> str:
    """Remove caracteres inválidos do nome da turma para uso no filename (AC-06)."""
    sanitizado = re.sub(r"[^a-zA-Z0-9_\-]", "_", turma)
    return sanitizado or "turma"


def gerar_nome_arquivo(turma: str, data: datetime | None = None) -> str:
    """Gera nome do arquivo `notas_{turma}_{data}.xlsx` (AC-06)."""
    turma_safe = sanitizar_nome_arquivo(turma) if turma else "turma"
    data_str = (data or datetime.now()).strftime("%Y%m%d")
    return f"notas_{turma_safe}_{data_str}.xlsx"


def _tipo_alerta(tem_plagio: bool, tem_grupo: bool) -> str:
    if tem_plagio and tem_grupo:
        return "Ambos"
    if tem_plagio:
        return "Plágio"
    if tem_grupo:
        return "Grupo"
    return ""


def _detalhes_plagio(ra: str, alertas_plagio: list[ParPlagio]) -> str:
    pares = [p for p in alertas_plagio if p.aluno_a == ra or p.aluno_b == ra]
    pares_relevantes = [p for p in pares if p.similaridade >= _LIMIAR_PLAGIO_MODERADO]
    if not pares_relevantes:
        return ""
    partes = []
    for p in sorted(pares_relevantes, key=lambda x: x.similaridade, reverse=True):
        ra_par = p.aluno_b if p.aluno_a == ra else p.aluno_a
        partes.append(f"RA {ra_par} ({p.similaridade:.0%})")
    return "; ".join(partes)


def _detalhes_grupo(ra: str, grupos_candidatos: list[GrupoCandidato]) -> str:
    grupos = [g for g in grupos_candidatos if ra in g.membros]
    if not grupos:
        return ""
    partes = []
    for g in grupos:
        outros = [m for m in sorted(g.membros) if m != ra]
        partes.append(f"Grupo com {', '.join(outros)}")
    return "; ".join(partes)


def _acao_para_label(acao: str) -> str:
    mapa = {
        "aprovado": "Aprovado",
        "aprovado_lote": "Aprovado",
        "editado": "Editado",
        "revisao_manual": "Revisão Manual",
    }
    return mapa.get(acao, acao)


def _estilizar_cabecalho(ws: Any, row: int) -> None:
    for cell in ws[row]:
        cell.font = _FONT_CABECALHO
        cell.fill = _FILL_CABECALHO
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autoajustar_colunas(ws: Any) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def construir_workbook(
    fichas_calibradas: list[FichaCorrecao],
    alertas_plagio: list[ParPlagio],
    grupos_candidatos: list[GrupoCandidato],
    decisoes: dict[str, Any],
    nomes_por_ra: dict[str, str],
    metadata: dict[str, Any],
) -> Workbook:
    """Constrói workbook openpyxl com abas Notas e Alertas (AC-02).

    Processamento 100% local — sem chamadas externas (AC-08).
    """
    wb = Workbook()

    _construir_aba_notas(
        wb, fichas_calibradas, alertas_plagio, grupos_candidatos, decisoes, nomes_por_ra, metadata
    )
    _construir_aba_alertas(wb, fichas_calibradas, alertas_plagio, grupos_candidatos, decisoes)

    return wb


def _construir_aba_notas(
    wb: Workbook,
    fichas_calibradas: list[FichaCorrecao],
    alertas_plagio: list[ParPlagio],
    grupos_candidatos: list[GrupoCandidato],
    decisoes: dict[str, Any],
    nomes_por_ra: dict[str, str],
    metadata: dict[str, Any],
) -> None:
    """Preenche aba 'Notas' (AC-03)."""
    ws = wb.active
    ws.title = "Notas"

    turma_meta = str(metadata.get("turma", ""))

    cabecalhos = [
        "RA",
        "Nome",
        "Turma",
        "Nota Bruta",
        "Nota Calibrada",
        "Nota Final",
        "Status",
        "Revisão Manual",
        "Observações PA",
    ]
    ws.append(cabecalhos)
    _estilizar_cabecalho(ws, 1)

    fichas_por_ra = {f.ra: f for f in fichas_calibradas}
    ras_ordenados = sorted(fichas_por_ra.keys())

    for ra in ras_ordenados:
        ficha = fichas_por_ra[ra]
        nota_calibrada = float(ficha.nota_a1) if ficha.nota_a1 is not None else 0.0
        nota_bruta = float(ficha.nota_a2) if ficha.nota_a2 is not None else nota_calibrada

        decisao = decisoes.get(ra, {})
        nota_final = float(decisao.get("nota_final", nota_calibrada))
        acao = decisao.get("acao", "aprovado")
        revisao_manual = acao == "revisao_manual"
        observacoes = str(decisao.get("observacao", ""))

        # Status badge
        status = "Normal"
        if nota_calibrada > 9.0:
            status = "Top 10%"
        elif nota_calibrada == 9.0 and nota_bruta > 9.0:
            status = "Cap 9"

        nome = str(nomes_por_ra.get(ra, ""))

        # turma por aluno (vem da planilha de upload); fallback para metadata
        turma_aluno = turma_meta

        row_data = [
            str(ra),
            nome,
            turma_aluno,
            round(nota_bruta, 1),
            round(nota_calibrada, 1),
            round(nota_final, 1),
            status,
            revisao_manual,
            observacoes,
        ]
        ws.append(row_data)

        # AC-05: linha amarela para revisão manual
        if revisao_manual:
            ultima_linha = ws.max_row
            for cell in ws[ultima_linha]:
                cell.fill = _FILL_REVISAO_MANUAL

    _autoajustar_colunas(ws)


def _construir_aba_alertas(
    wb: Workbook,
    fichas_calibradas: list[FichaCorrecao],
    alertas_plagio: list[ParPlagio],
    grupos_candidatos: list[GrupoCandidato],
    decisoes: dict[str, Any],
) -> None:
    """Preenche aba 'Alertas' (AC-04)."""
    ws = wb.create_sheet(title="Alertas")

    cabecalhos = ["RA", "Tipo de Alerta", "Detalhes", "Decisão do PA"]
    ws.append(cabecalhos)
    _estilizar_cabecalho(ws, 1)

    ras_com_alerta: set[str] = set()
    for p in alertas_plagio:
        if p.similaridade >= _LIMIAR_PLAGIO_MODERADO:
            ras_com_alerta.add(p.aluno_a)
            ras_com_alerta.add(p.aluno_b)
    for g in grupos_candidatos:
        for m in g.membros:
            ras_com_alerta.add(m)

    for ra in sorted(ras_com_alerta):
        tem_plagio = any(
            (p.aluno_a == ra or p.aluno_b == ra) and p.similaridade >= _LIMIAR_PLAGIO_MODERADO
            for p in alertas_plagio
        )
        tem_grupo = any(ra in g.membros for g in grupos_candidatos)

        tipo = _tipo_alerta(tem_plagio, tem_grupo)

        detalhes_parts = []
        if tem_plagio:
            d = _detalhes_plagio(ra, alertas_plagio)
            if d:
                detalhes_parts.append(d)
        if tem_grupo:
            d = _detalhes_grupo(ra, grupos_candidatos)
            if d:
                detalhes_parts.append(d)
        detalhes = " | ".join(detalhes_parts)

        decisao = decisoes.get(ra, {})
        acao_label = _acao_para_label(decisao.get("acao", ""))

        ws.append([str(ra), tipo, detalhes, acao_label])

    _autoajustar_colunas(ws)


def workbook_para_bytes(wb: Workbook) -> bytes:
    """Serializa workbook para bytes para uso com st.download_button."""
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
