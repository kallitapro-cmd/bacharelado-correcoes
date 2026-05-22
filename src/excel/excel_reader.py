"""Leitor da planilha da turma — Aba 1 (Story 1.10)."""

from __future__ import annotations

from typing import TYPE_CHECKING

import openpyxl
from pydantic import ValidationError

if TYPE_CHECKING:
    from pathlib import Path

from src.matching.ra_normalizer import normalize_ra
from src.models.schemas import Aluno


def _find_col(header: dict[str, int], candidates: list[str]) -> int | None:
    """Retorna o índice da primeira coluna que case com algum candidato."""
    for name in candidates:
        if name in header:
            return header[name]
    return None


def read_aba1(file_path: str | Path) -> list[Aluno]:
    """Lê a Aba 1 (Alunos) da planilha do PA.

    Abre o workbook em modo somente-leitura (read_only=True) — impossibilita
    fisicamente qualquer escrita na planilha (recomendação do Pedro Valério /
    auditoria pré-QA).

    Detecta colunas por nome (case-insensitive, strip), normaliza todos os RAs
    via normalize_ra() e retorna lista de Aluno.

    RAs de Camada 3 (ex.: 0000xxxxxx) que não satisfazem o pattern 11 dígitos
    do schema Aluno são incluídos via model_construct — preservando o dado para
    revisão manual sem levantar ValidationError.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    ws = wb.worksheets[0]

    header = {
        cell.value.strip().lower(): idx
        for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))
        if cell.value
    }

    col_ra = _find_col(header, ["ra", "registro acadêmico", "código"])
    col_nome = _find_col(header, ["nome completo", "nome", "aluno"])
    col_email = _find_col(header, ["e-mail", "email", "e_mail"])
    col_telefone = _find_col(header, ["telefone", "tel", "celular"])

    if col_ra is None:
        raise ValueError("Coluna RA não encontrada na Aba 1")

    alunos: list[Aluno] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        ra_bruto = str(row[col_ra]).strip() if row[col_ra] is not None else ""
        if not ra_bruto:
            continue

        ra_normalizado = normalize_ra(ra_bruto)
        nome = str(row[col_nome]) if col_nome is not None and row[col_nome] else ""
        email = str(row[col_email]) if col_email is not None and row[col_email] else None
        telefone = (
            str(row[col_telefone]) if col_telefone is not None and row[col_telefone] else None
        )

        try:
            aluno = Aluno(ra=ra_normalizado, nome=nome, email=email, telefone=telefone)
        except ValidationError:
            aluno = Aluno.model_construct(
                ra=ra_normalizado, nome=nome, email=email, telefone=telefone
            )

        alunos.append(aluno)

    wb.close()
    return alunos
